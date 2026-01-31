"""Excel generation utility using openpyxl."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Dict, List, Any
from datetime import datetime


class RFPExcelGenerator:
    """Generate formatted Excel workbook from RFP analysis."""
    
    def __init__(self):
        """Initialize Excel generator with styling."""
        self.header_fill = PatternFill(
            start_color="2563EB",
            end_color="2563EB",
            fill_type="solid"
        )
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_workbook(self, next_steps: List[str], 
                         alignment_questions: List[str],
                         rfp_filename: str = "RFP") -> BytesIO:
        """
        Create 2-tab Excel workbook with next steps and alignment questions.
        
        Args:
            next_steps: List of action items for proposal team
            alignment_questions: List of strategic questions
            rfp_filename: Original RFP filename for reference
        
        Returns:
            BytesIO buffer containing Excel file
        """
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Tab 1: Next Steps
        ws_steps = wb.create_sheet("Next Steps")
        self._create_next_steps_sheet(ws_steps, next_steps, rfp_filename)
        
        # Tab 2: Alignment Questions
        ws_questions = wb.create_sheet("Alignment Questions")
        self._create_alignment_questions_sheet(ws_questions, alignment_questions, rfp_filename)
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _create_next_steps_sheet(self, ws, next_steps: List[str], 
                                 rfp_filename: str):
        """
        Create the Next Steps worksheet.
        
        Args:
            ws: Worksheet object
            next_steps: List of action items
            rfp_filename: Original filename
        """
        # Title
        ws['A1'] = "Next Steps for Proposal Team"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:C1')
        
        # Metadata
        ws['A2'] = f"RFP: {rfp_filename}"
        ws['A2'].font = Font(italic=True, size=10)
        ws.merge_cells('A2:C2')
        
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A3'].font = Font(italic=True, size=10)
        ws.merge_cells('A3:C3')
        
        # Header row
        row = 5
        ws[f'A{row}'] = "#"
        ws[f'B{row}'] = "Action Item"
        ws[f'C{row}'] = "Status"
        
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{row}']
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data rows
        row += 1
        for idx, step in enumerate(next_steps, 1):
            ws[f'A{row}'] = idx
            ws[f'B{row}'] = step
            ws[f'C{row}'] = "Pending"
            
            # Apply formatting
            for col in ['A', 'B', 'C']:
                cell = ws[f'{col}{row}']
                cell.border = self.border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='top')
            
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 70
        ws.column_dimensions['C'].width = 15
    
    def _create_alignment_questions_sheet(self, ws, questions: List[str],
                                          rfp_filename: str):
        """
        Create the Alignment Questions worksheet.
        
        Args:
            ws: Worksheet object
            questions: List of strategic questions
            rfp_filename: Original filename
        """
        # Title
        ws['A1'] = "Strategic Alignment Questions"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:C1')
        
        # Metadata
        ws['A2'] = f"RFP: {rfp_filename}"
        ws['A2'].font = Font(italic=True, size=10)
        ws.merge_cells('A2:C2')
        
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A3'].font = Font(italic=True, size=10)
        ws.merge_cells('A3:C3')
        
        # Header row
        row = 5
        ws[f'A{row}'] = "#"
        ws[f'B{row}'] = "Question"
        ws[f'C{row}'] = "Answer / Notes"
        
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{row}']
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data rows
        row += 1
        for idx, question in enumerate(questions, 1):
            ws[f'A{row}'] = idx
            ws[f'B{row}'] = question
            ws[f'C{row}'] = ""  # Empty for team to fill in
            
            # Apply formatting
            for col in ['A', 'B', 'C']:
                cell = ws[f'{col}{row}']
                cell.border = self.border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='top')
            
            # Add light background for answer column
            ws[f'C{row}'].fill = PatternFill(
                start_color="F3F4F6",
                end_color="F3F4F6",
                fill_type="solid"
            )
            
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 50
